#!/usr/bin/env python3
"""
NEXO Headhunter RFQ — Procesador manual + COT generator
Uso: python3 headhunter_rfq.py "texto del RFQ"
       echo "texto" | python3 headhunter_rfq.py
"""
import sys, os, re, json, httpx
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
from openpyxl.utils import get_column_letter
import time, urllib.parse
from bs4 import BeautifulSoup

# ── Paths ──────────────────────────────────────────────────────────────────────
DB_PATH  = Path.home() / "MICSA-Brain" / "RFQs" / "MICSA_RFQ_DATABASE.xlsx"
COT_DIR  = Path.home() / "MICSA-Brain" / "RFQs" / "COTs"
COT_DIR.mkdir(parents=True, exist_ok=True)

# ── Telegram ───────────────────────────────────────────────────────────────────
TELEGRAM_TOKEN = "8621290655:AAG3QKnb5JOCvgqL58uGSm5v2ezqQw29J80"
JORDAN_ID      = "1579401409"

# ── Styles ─────────────────────────────────────────────────────────────────────
NAVY   = "0A1628"
GOLD   = "F5B800"
GREEN  = "008000"
BLUE   = "0000FF"
WHITE  = "FFFFFF"
LGRAY  = "F2F2F2"

thin   = Side(border_style="thin", color="CCCCCC")
border = Border(top=thin, bottom=thin, left=thin, right=thin)


def hfont(bold=True, color=WHITE, size=11):
    return Font(name="Calibri", size=size, bold=bold, color=color)


def hfill(color):
    return PatternFill("solid", fgColor=color)


# ── 1. PARSE RFQ ───────────────────────────────────────────────────────────────
def parse_rfq(texto: str) -> dict:
    t = texto.strip()
    urgencia = "Media"
    for kw in ["urgente", "inmediata", "para hoy", "necesito ya", "inmediato", "asap"]:
        if kw in t.lower():
            urgencia = "Inmediata"
            break
    for kw in ["esta semana", "pronto", "lo antes posible"]:
        if kw in t.lower():
            urgencia = "Alta"
            break

    # cantidad
    qty_match = re.search(r"(\d+)\s*(pz|pzs|piezas?|unidades?|u\b|tons?|kg|lt|m\b|metros?|sets?)", t, re.I)
    cantidad = qty_match.group(0) if qty_match else "Por confirmar"

    # ubicacion
    ciudades = ["cdmx", "monterrey", "mty", "guadalajara", "gdl", "querétaro", "queretaro",
                "tlaxcala", "veracruz", "aguascalientes", "saltillo", "tijuana", "san luis", "slp",
                "puebla", "leon", "silao", "toluca", "chihuahua"]
    ubicacion = "No declarada"
    for c in ciudades:
        if c in t.lower():
            ubicacion = c.title()
            break

    # marca / modelo — heuristica simple
    marca_match = re.search(r"\b(siemens|endress|hauser|carrier|mac|sew|mitutoyo|rexroth|wago|weidmuller|zebra|allen[-\s]?bradley|schneider|honeywell|parker|festo|smc|abb|danfoss|emerson)\b", t, re.I)
    marca = marca_match.group(1).title() if marca_match else "Por confirmar"

    # modelo / numero de parte
    modelo_match = re.search(r"\b([A-Z0-9][-A-Z0-9/]{3,})\b", t)
    modelo = modelo_match.group(1) if modelo_match else ""

    # categoria
    cat = detectar_categoria(t)
    match = detectar_match(t, cat)

    return {
        "texto_original": t,
        "producto": extraer_producto(t),
        "marca": marca,
        "modelo": modelo,
        "cantidad": cantidad,
        "ubicacion": ubicacion,
        "urgencia": urgencia,
        "categoria": cat,
        "match_micsa": match,
        "tipo": "Por determinar",
    }


def extraer_producto(t):
    for pat in [r"(?:rfq[:\s.\/]*|busco|requiero|necesito)\s+(.{5,80}?)(?:\s+en\s|\s+para\s|,|\.|$)", ]:
        m = re.search(pat, t, re.I)
        if m:
            return m.group(1).strip()
    return t[:80].strip()


def detectar_categoria(t):
    t = t.lower()
    cats = {
        "EPP": ["epp", "casco", "guante", "botas seguridad", "lente", "arnés", "arnes", "zapatos seguridad", "uniforme"],
        "MRO Eléctrico": ["cable", "breaker", "tablero", "interruptor", "transformador", "medidor", "contacto eléctrico"],
        "MRO Mecánico": ["rodamiento", "chumacera", "banda", "cadena", "sello", "empaque", "valvula", "válvula", "reductor", "motor"],
        "MRO Neumática": ["valvula mac", "actuador", "cilindro neumático", "frl", "manguera neumatica"],
        "Instrumentación": ["medidor de flujo", "transmisor", "sensor", "termómetro", "presostato", "flowmeter", "coriolis", "endress"],
        "HVAC/Instalaciones": ["carrier", "hvac", "aire acondicionado", "ducto", "climatizacion"],
        "Automatización": ["plc", "siemens", "variador", "hmi", "scada", "encoder", "servo"],
        "Mantenimiento Industrial": ["vulcanizado", "mantenimiento", "servicio técnico", "reparacion", "overhaul"],
        "Maniobras/Izajes": ["grua", "maniobra", "izaje", "montacargas", "pluma"],
        "Servicios HSE": ["segurista", "supervisor seguridad", "hse", "residuos", "recoleccion residuos"],
        "Transporte": ["transporte", "flete", "logística", "paquetería"],
        "Limpieza Industrial": ["trapo industrial", "absorbente", "limpieza"],
        "TI/Software": ["laptop", "servidor", "software", "licencia", "antivirus"],
    }
    for cat, kws in cats.items():
        for kw in kws:
            if kw in t:
                return cat
    return "MRO General"


def detectar_match(t, cat):
    alto = ["maniobras", "izajes", "grua", "montaje", "mantenimiento industrial", "transporte carga"]
    medio = ["EPP", "MRO Eléctrico", "MRO Mecánico", "MRO Neumática", "Instrumentación", "HVAC", "Automatización"]
    for kw in alto:
        if kw in t.lower():
            return "Alto"
    if any(cat.startswith(m) for m in medio):
        return "Medio"
    return "Bajo"


# ── 2. WEB SEARCH ─────────────────────────────────────────────────────────────
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "es-MX,es;q=0.9,en;q=0.8",
}


def google_search(query: str, num: int = 5) -> list[dict]:
    """Search via DuckDuckGo HTML (no API key, no JS required)."""
    q = urllib.parse.quote_plus(query)
    # Intentar DuckDuckGo HTML primero, luego Bing como fallback
    for url in [
        f"https://html.duckduckgo.com/html/?q={q}",
        f"https://www.bing.com/search?q={q}&count={num}&mkt=es-MX",
    ]:
        try:
            r = httpx.get(url, headers=HEADERS, timeout=15, follow_redirects=True)
            soup = BeautifulSoup(r.text, "html.parser")
            results = []

            # DuckDuckGo HTML
            for item in soup.select("div.result, div.web-result")[:num]:
                a = item.select_one("a.result__a, h2 a")
                snippet_el = item.select_one("a.result__snippet, div.b_caption p")
                if not a:
                    continue
                href = a.get("href", "")
                # DDG redirect
                if "duckduckgo.com/l/" in href:
                    m = re.search(r"uddg=([^&]+)", href)
                    href = urllib.parse.unquote(m.group(1)) if m else href
                results.append({
                    "title": a.get_text()[:100],
                    "snippet": snippet_el.get_text()[:300] if snippet_el else "",
                    "url": href,
                })

            # Bing fallback
            if not results:
                for li in soup.select("li.b_algo")[:num]:
                    a = li.select_one("h2 a")
                    snippet_el = li.select_one("div.b_caption p, p.b_algoSlug")
                    if not a:
                        continue
                    results.append({
                        "title": a.get_text()[:100],
                        "snippet": snippet_el.get_text()[:300] if snippet_el else "",
                        "url": a.get("href", ""),
                    })

            if results:
                return results
            time.sleep(1)
        except Exception as e:
            print(f"      [warn] search ({url[:30]}...): {e}")
            time.sleep(1)
    return []


def buscar_ficha(rfq: dict) -> list[dict]:
    if rfq["categoria"] in ["Maniobras/Izajes", "Servicios HSE", "Transporte", "Mantenimiento Industrial"]:
        return []

    q1 = f"{rfq['marca']} {rfq['modelo']} datasheet filetype:pdf"
    q2 = f"{rfq['producto']} {rfq['marca']} especificaciones técnicas"
    results = google_search(q1, 4) or google_search(q2, 4)
    specs = []
    for r in results[:3]:
        specs.append({"title": r["title"], "snippet": r["snippet"][:300], "url": r["url"]})
    return specs


def buscar_proveedor_precio(rfq: dict) -> list[dict]:
    q = f"{rfq['marca']} {rfq['modelo']} {rfq['producto']} precio cotización proveedor México"
    results = google_search(q, 8)

    proveedores = []
    precio_regex = re.compile(r"(?:USD?\s*|MXN?\s*|\$\s*)([\d,]+(?:\.\d{1,2})?)", re.I)

    for r in results[:6]:
        snippet = r["snippet"]
        url = r["url"]
        title = r["title"]
        precios = precio_regex.findall(snippet + " " + title)

        precio_usd = None
        for p in precios:
            val = float(p.replace(",", ""))
            if val < 1:
                continue
            if "mxn" in (snippet + title).lower() or val > 5000:
                precio_usd = round(val / 17.5, 2)
            else:
                precio_usd = round(val, 2)
            break

        nombre = title[:60] or (url.split("/")[2] if "/" in url else url[:60])
        proveedores.append({
            "nombre": nombre,
            "url": url,
            "precio_usd": precio_usd,
            "snippet": snippet[:200],
        })

    con_precio = [p for p in proveedores if p["precio_usd"]]
    return con_precio or proveedores[:3]


# ── 3. GENERAR COT EXCEL ──────────────────────────────────────────────────────
def get_next_cot_num() -> int:
    existing = sorted(COT_DIR.glob("COT-MICSA-2026-*.xlsx"))
    if not existing:
        return 1
    last = existing[-1].stem  # COT-MICSA-2026-0012
    return int(last.split("-")[-1]) + 1


def generar_cot(rfq: dict, specs: list, proveedores: list, fecha: str, rfq_id: str) -> Path:
    num = get_next_cot_num()
    filename = COT_DIR / f"COT-MICSA-2026-{num:04d}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"

    # columnas
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    def cell(row, col, value, bold=False, color="000000", fill_color=None, align="left", fmt=None, italic=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Calibri", size=11, bold=bold, color=color, italic=italic)
        if fill_color:
            c.fill = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        c.border = border
        if fmt:
            c.number_format = fmt
        return c

    def merge_row(row, value, bold=True, color=WHITE, fill_color=NAVY, align="center"):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=value)
        c.font = Font(name="Calibri", size=12, bold=bold, color=color)
        c.fill = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(horizontal=align, vertical="center")
        ws.row_dimensions[row].height = 20
        return c

    # ── A. HEADER ──────────────────────────────────────────────────────────────
    merge_row(1, "GRUPO MICSA — COTIZACIÓN DE SUMINISTRO")
    merge_row(2, f"COT-MICSA-2026-{num:04d}  |  {fecha}  |  Validez: 15 días", bold=False, fill_color=GOLD, color="000000")

    ws.row_dimensions[3].height = 6
    cell(4, 1, "RFQ Origen", bold=True)
    cell(4, 2, rfq_id)
    cell(4, 3, "Categoría", bold=True)
    cell(4, 4, rfq["categoria"])

    cell(5, 1, "Ubicación", bold=True)
    cell(5, 2, rfq["ubicacion"])
    cell(5, 3, "Urgencia", bold=True)
    cell(5, 4, rfq["urgencia"], color="CC0000" if rfq["urgencia"] == "Inmediata" else "000000")

    # ── B. PRODUCTO ────────────────────────────────────────────────────────────
    ws.row_dimensions[6].height = 6
    merge_row(7, "DESCRIPCIÓN DEL REQUERIMIENTO", fill_color=NAVY)

    cell(8, 1, "Producto / Servicio", bold=True)
    cell(8, 2, rfq["producto"], color=BLUE)
    cell(8, 3, "Marca", bold=True)
    cell(8, 4, rfq["marca"], color=BLUE)

    cell(9, 1, "Modelo / P/N", bold=True)
    cell(9, 2, rfq["modelo"] or "Por confirmar", color=BLUE)
    cell(9, 3, "Cantidad", bold=True)
    cell(9, 4, rfq["cantidad"], color=BLUE)

    cell(10, 1, "Texto original del RFQ", bold=True)
    ws.merge_cells(f"B10:D10")
    c = ws.cell(row=10, column=2, value=rfq["texto_original"][:300])
    c.font = Font(name="Calibri", size=9, italic=True, color="555555")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[10].height = 50

    # ── C. FICHA TÉCNICA ───────────────────────────────────────────────────────
    ws.row_dimensions[11].height = 6
    merge_row(12, "FICHA TÉCNICA (REFERENCIA)", fill_color=NAVY)

    if specs:
        row = 13
        for i, s in enumerate(specs[:4]):
            cell(row, 1, f"Fuente {i+1}", bold=True)
            cell(row, 2, s["title"][:60])
            ws.merge_cells(f"C{row}:D{row}")
            c = ws.cell(row=row, column=3, value=s["url"])
            c.font = Font(name="Calibri", size=9, color="0563C1", italic=True)
            c.alignment = Alignment(horizontal="left")
            row += 1
            ws.merge_cells(f"A{row}:D{row}")
            c = ws.cell(row=row, column=1, value=s["snippet"])
            c.font = Font(name="Calibri", size=9, color="444444")
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 40
            row += 1
        next_section = row
    else:
        cell(13, 1, "Tipo", bold=True)
        cell(13, 2, "Servicio — ficha técnica no aplica", italic=True)
        next_section = 14

    # ── D. PROVEEDORES ─────────────────────────────────────────────────────────
    ws.row_dimensions[next_section].height = 6
    pr = next_section + 1
    merge_row(pr, "PROVEEDORES IDENTIFICADOS", fill_color=NAVY)
    pr += 1

    headers_p = ["Proveedor", "Precio USD (ref)", "Fuente"]
    for col, h in enumerate(headers_p, 1):
        cell(pr, col, h, bold=True, fill_color=GOLD, color="000000", align="center")
    pr += 1

    mejor_precio = None
    for i, prov in enumerate(proveedores[:4]):
        fill = "E8F5E9" if i == 0 else None
        cell(pr, 1, prov["nombre"][:60], fill_color=fill)
        precio_cell = ws.cell(row=pr, column=2)
        if prov["precio_usd"]:
            precio_cell.value = prov["precio_usd"]
            precio_cell.number_format = '"$"#,##0.00'
            precio_cell.font = Font(name="Calibri", size=11, color=GREEN if i == 0 else "000000", bold=(i==0))
            if i == 0:
                mejor_precio = prov["precio_usd"]
        else:
            precio_cell.value = "Por cotizar"
            precio_cell.font = Font(name="Calibri", size=11, color="888888")
        precio_cell.fill = PatternFill("solid", fgColor=fill) if fill else PatternFill()
        precio_cell.border = border
        precio_cell.alignment = Alignment(horizontal="right")
        c = ws.cell(row=pr, column=3, value=prov["url"])
        c.font = Font(name="Calibri", size=9, color="0563C1", italic=True)
        c.border = border
        c.alignment = Alignment(wrap_text=True)
        pr += 1

    if not proveedores:
        cell(pr, 1, "Sin proveedores encontrados — cotizar manualmente")
        pr += 1

    # ── E. COTIZACIÓN MICSA ────────────────────────────────────────────────────
    ws.row_dimensions[pr].height = 6
    cq = pr + 1
    merge_row(cq, "COTIZACIÓN GRUPO MICSA", fill_color=NAVY)
    cq += 1

    # Cantidad input
    cell(cq, 1, "Cantidad", bold=True)
    qty_val = 1
    qty_match = re.search(r"(\d+)", rfq["cantidad"])
    if qty_match:
        qty_val = int(qty_match.group(1))
    qty_cell = ws.cell(row=cq, column=2, value=qty_val)
    qty_cell.font = Font(name="Calibri", size=11, color=BLUE, bold=True)
    qty_cell.border = border
    qty_ref = f"B{cq}"
    cq += 1

    # Precio unitario
    cell(cq, 1, "Precio Unit (ref)", bold=True)
    if mejor_precio:
        pu_cell = ws.cell(row=cq, column=2, value=mejor_precio)
        pu_cell.number_format = '"$"#,##0.00'
    else:
        pu_cell = ws.cell(row=cq, column=2, value=None)
        pu_cell.number_format = '"$"#,##0.00'
    pu_cell.font = Font(name="Calibri", size=11, color=BLUE, bold=True)
    pu_cell.border = border
    pu_ref = f"B{cq}"
    cq += 1

    # Subtotal
    cell(cq, 1, "Subtotal", bold=True)
    st_cell = ws.cell(row=cq, column=2)
    st_cell.value = f"={qty_ref}*{pu_ref}"
    st_cell.number_format = '"$"#,##0.00'
    st_cell.font = Font(name="Calibri", size=11)
    st_cell.border = border
    st_ref = f"B{cq}"
    cq += 1

    # Margen 30%
    cell(cq, 1, "Margen MICSA (30%)", bold=True)
    mg_cell = ws.cell(row=cq, column=2)
    mg_cell.value = f"={st_ref}*0.3"
    mg_cell.number_format = '"$"#,##0.00'
    mg_cell.font = Font(name="Calibri", size=11)
    mg_cell.border = border
    mg_ref = f"B{cq}"
    cq += 1

    # Total MICSA
    ws.merge_cells(f"A{cq}:A{cq}")
    total_label = ws.cell(row=cq, column=1, value="PRECIO FINAL MICSA")
    total_label.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    total_label.fill = PatternFill("solid", fgColor=NAVY)
    total_label.border = border
    total_label.alignment = Alignment(horizontal="center", vertical="center")

    total_cell = ws.cell(row=cq, column=2)
    total_cell.value = f"={st_ref}+{mg_ref}"
    total_cell.number_format = '"$"#,##0.00'
    total_cell.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    total_cell.fill = PatternFill("solid", fgColor=GREEN)
    total_cell.border = border
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[cq].height = 22
    total_ref = f"B{cq}"
    cq += 1

    # Condiciones
    ws.row_dimensions[cq].height = 6
    cq += 1
    merge_row(cq, "Anticipo: 50% | Saldo: contra entrega | Validez: 15 días", bold=False, fill_color=GOLD, color="000000")

    wb.save(filename)
    return filename, num, mejor_precio


# ── 4. APPEND EXCEL MAESTRO ───────────────────────────────────────────────────
def get_next_id(ws, col=1, prefix="RFQ") -> str:
    """Lee el último ID de la columna y suma 1."""
    ids = [ws.cell(row=r, column=col).value for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=col).value]
    if not ids:
        stamp = datetime.now().strftime("%y%m%d-%H%M")
        return f"{prefix}-{stamp}-1"
    last = str(ids[-1])
    parts = last.rsplit("-", 1)
    try:
        return f"{parts[0]}-{int(parts[1])+1}"
    except:
        stamp = datetime.now().strftime("%y%m%d-%H%M")
        return f"{prefix}-{stamp}-1"


def append_maestro(rfq: dict, cot_path: Path, cot_num: int, mejor_precio, rfq_id: str, fecha: str):
    wb = load_workbook(DB_PATH)

    # RFQs
    ws = wb["RFQs"]
    ws.append([
        rfq_id,
        rfq["producto"],
        rfq["marca"],
        rfq["modelo"],
        rfq["cantidad"],
        rfq["ubicacion"],
        rfq["urgencia"],
        rfq["categoria"],
        rfq["tipo"],
        rfq["match_micsa"],
        "COT GENERADO",
        fecha,
        "",   # Link Post
        "",   # Autor
        "",   # Empresa
    ])

    # COTs Generadas
    ws2 = wb["COTs Generadas"]
    total_est = round(mejor_precio * 1.3, 2) if mejor_precio else None
    ws2.append([
        f"COT-MICSA-2026-{cot_num:04d}",
        rfq_id,
        rfq["producto"],
        fecha,
        mejor_precio,
        total_est,
        str(cot_path),
    ])

    wb.save(DB_PATH)


# ── 5. TELEGRAM NOTIFY ────────────────────────────────────────────────────────
def telegram_send(mensaje: str, archivo: Path = None):
    base = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}"
    if archivo and archivo.exists():
        with open(archivo, "rb") as f:
            r = httpx.post(
                f"{base}/sendDocument",
                data={"chat_id": JORDAN_ID, "caption": mensaje[:1024], "parse_mode": "Markdown"},
                files={"document": (archivo.name, f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=30,
            )
    else:
        r = httpx.post(f"{base}/sendMessage", json={"chat_id": JORDAN_ID, "text": mensaje, "parse_mode": "Markdown"}, timeout=15)
    return r.status_code == 200


# ── MAIN ───────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) > 1:
        texto = " ".join(sys.argv[1:])
    elif not sys.stdin.isatty():
        texto = sys.stdin.read()
    else:
        print("Uso: python3 headhunter_rfq.py \"texto del RFQ\"")
        sys.exit(1)

    fecha = datetime.now().strftime("%Y-%m-%d %H:%M")
    stamp = datetime.now().strftime("%y%m%d-%H%M")

    print(f"[1/5] Parseando RFQ...")
    rfq = parse_rfq(texto)

    # Asignar ID
    wb = load_workbook(DB_PATH)
    rfq_id = get_next_id(wb["RFQs"], col=1, prefix="RFQ")
    wb.close()

    print(f"      Producto: {rfq['producto']}")
    print(f"      Categoría: {rfq['categoria']} | Match MICSA: {rfq['match_micsa']} | Urgencia: {rfq['urgencia']}")

    print(f"[2/5] Buscando ficha técnica...")
    specs = buscar_ficha(rfq)
    print(f"      {len(specs)} fuentes encontradas")

    print(f"[3/5] Buscando proveedor y precio...")
    proveedores = buscar_proveedor_precio(rfq)
    mejor = next((p["precio_usd"] for p in proveedores if p["precio_usd"]), None)
    print(f"      {len(proveedores)} proveedores | Mejor precio ref: {'${:,.2f} USD'.format(mejor) if mejor else 'No encontrado'}")

    print(f"[4/5] Generando COT Excel...")
    cot_path, cot_num, mejor_precio = generar_cot(rfq, specs, proveedores, fecha, rfq_id)
    print(f"      Archivo: {cot_path.name}")

    print(f"[5/5] Guardando en base de datos y enviando por Telegram...")
    append_maestro(rfq, cot_path, cot_num, mejor_precio, rfq_id, fecha)

    total_micsa = round(mejor_precio * 1.3, 2) if mejor_precio else None
    msg = (
        f"*NEXO Headhunter — Nueva COT*\n\n"
        f"*RFQ:* `{rfq_id}`\n"
        f"*Producto:* {rfq['producto']}\n"
        f"*Marca:* {rfq['marca']} | *Modelo:* {rfq['modelo'] or '-'}\n"
        f"*Categoría:* {rfq['categoria']}\n"
        f"*Ubicación:* {rfq['ubicacion']} | *Urgencia:* {rfq['urgencia']}\n"
        f"*Match MICSA:* {rfq['match_micsa']}\n\n"
        f"*Mejor precio ref:* {'${:,.2f} USD'.format(mejor_precio) if mejor_precio else 'Por confirmar'}\n"
        f"*Precio MICSA +30%:* {'${:,.2f} USD'.format(total_micsa) if total_micsa else 'Por confirmar'}\n\n"
        f"Archivo adjunto: `{cot_path.name}`"
    )
    ok = telegram_send(msg, cot_path)

    print()
    print("=" * 60)
    print(f"COT generada: {cot_path}")
    print(f"RFQ ID: {rfq_id}")
    if mejor_precio:
        print(f"Precio ref: ${mejor_precio:,.2f} USD → con 30% MICSA: ${total_micsa:,.2f} USD")
    print(f"Telegram: {'OK' if ok else 'FALLO (revisa token)'}")
    print("=" * 60)


if __name__ == "__main__":
    main()
